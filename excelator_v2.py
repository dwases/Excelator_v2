from openpyxl import load_workbook
from openpyxl import Workbook
#import re

new_wb = Workbook()
old_wb = load_workbook("C61 2014-2021.xlsx")
new_ws1 = new_wb.active
new_ws1.title = "sheet 1"
old_ws1 = old_wb.active

old_colA = old_ws1['A']

new_colA = new_ws1['A'] #nazwisko
new_colA = new_ws1['B'] #imie
new_colA = new_ws1['C'] #data urodzenia

for i in range(0, 2056):
    try:
        lista = old_colA[i].value.split(" ")
        #print(lista)
        if len(lista) <= 1:
            pass
        else:
            new_ws1['A{}'.format(i)].value = lista[0]   #nazwisko
            new_ws1['B{}'.format(i)].value = lista[1]   #imie
            new_ws1['C{}'.format(i)].value = lista[2]   #data urodzenia
            new_ws1['C{}'.format(i)].value = new_ws1['C{}'.format(i)].value[3:]
            print([lista[0], lista[1], lista[2], new_ws1['C{}'.format(i)].value[:]])
    except:
        pass

new_wb.save(filename='dokument dane.xlsx')